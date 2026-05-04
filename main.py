# -*- coding: utf-8 -*-
"""
上画管理系统 - 安卓适配版
功能：小区楼栋单元电梯框位的筛选、状态标记、数据修改
"""
import sys
import pandas as pd
import openpyxl
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton,
    QHBoxLayout, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QFileDialog, QLabel, QMessageBox, QInputDialog,
    QListWidget, QListWidgetItem, QDialog
)
from PySide6.QtGui import QColor
from PySide6.QtCore import Qt, QPoint, Slot

# 安卓平台专用导入
is_android = sys.platform == 'android'
if is_android:
    from PySide6.QtAndroid import QtAndroid


class InfoManagerWindow(QMainWindow):
    """
    主窗口类
    负责整个程序的界面布局、事件处理、数据管理
    自动适配PC和安卓平台
    """

    def __init__(self):
        """初始化主窗口"""
        super().__init__()
        # ==================== 平台判断与基础设置 ====================
        self.is_android = is_android
        self.setWindowTitle("上画管理系统")

        # 安卓平台：全屏+竖屏锁定
        if self.is_android:
            # 安卓上全屏显示，适配手机屏幕
            self.showMaximized()
            # 锁定竖屏，避免横屏布局错乱
            QtAndroid.setOrientation(QtAndroid.ScreenOrientation.Portrait)
            # 安卓触摸适配的尺寸参数
            self.btn_fixed_width = 90  # 按钮宽度，触摸要更大
            self.btn_padding = 12  # 按钮内边距，放大点击区域
            self.table_row_height = 48  # 表格行高，触摸更容易点
            self.popup_item_height = 48  # 弹窗选项高度，触摸不点错
            self.popup_width = 280  # 弹窗宽度，手机上更宽
            self.font_size = 12  # 字体大小，手机上更大更清楚
            self.default_file_dir = '/sdcard/Download/'  # 安卓默认文件目录，用户的Excel一般存在这里
        else:
            # PC平台：保持原来的尺寸
            self.resize(300, 720)
            self.btn_fixed_width = 65
            self.btn_padding = 8
            self.table_row_height = 32
            self.popup_item_height = 28
            self.popup_width = 210
            self.font_size = 10
            self.default_file_dir = ''

        # 全局样式，根据平台调整字体
        self.setStyleSheet(f"""
            QMain{{background-color: #eef2f7;}}
            QLabel{{font-size: {self.font_size}pt;color: #2e3e45;}}
            QComboBox QAbstractScrollArea {{ border: none; }}
            QComboBox QScrollBar:vertical {{ width: 2px; }}
            QComboBox QScrollBar:horizontal {{ height: 2px; }}
        """)

        # ==================== 成员变量 ====================
        self.file_path = ""  # Excel文件路径
        self.df_original = None  # 原始完整数据（未筛选的全部数据）
        self.df_display = None  # 当前筛选后显示的数据
        self.row_styles = {}  # 行字体颜色样式：key=行唯一标识，value=颜色(green/red/black)
        # 筛选条件存储
        self.filter_condition = {
            "area": "全部",  # 小区筛选条件
            "build": "全部",  # 楼栋筛选条件
            "unit": "全部",  # 单元筛选条件
            "elevator": "全部",  # 电梯筛选条件
            "box": "全部"  # 框位筛选条件
        }
        # 权限申请回调用的临时变量
        self._pending_import = False

        # ==================== 界面布局 ====================
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(12)
        # 安卓上调整边距，避免太靠边
        margin = 16 if self.is_android else 20
        main_layout.setContentsMargins(margin, margin, margin, margin)

        # --- 顶部按钮：导入、另存为 ---
        top_btn_layout = QHBoxLayout()
        # 导入按钮
        self.import_btn = QPushButton("📁 打开")
        self.import_btn.setFixedWidth(self.btn_fixed_width)
        self.import_btn.setStyleSheet(f"""
            QPushButton{{
                background-color: #409eff;
                color: white;
                border: none;
                border-radius: 6px;
                padding: {self.btn_padding}px 0px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color: #66b1ff;}}
            QPushButton:pressed{{background-color: #337ecc;}}
        """)
        # 另存为按钮
        self.save_btn = QPushButton("💾 另存")
        self.save_btn.setFixedWidth(self.btn_fixed_width)
        self.save_btn.setStyleSheet(f"""
            QPushButton{{
                background-color: #00923A;
                color: white;
                border: none;
                border-radius: 6px;
                padding: {self.btn_padding}px 0px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color: #85ce61;}}
            QPushButton:pressed{{background-color: #529b2e;}}
        """)
        top_btn_layout.addWidget(self.import_btn)
        top_btn_layout.addWidget(self.save_btn)
        top_btn_layout.addStretch()
        main_layout.addLayout(top_btn_layout)

        # --- 统计标签：总数/已完成/未完成/异常 ---
        stats_layout = QHBoxLayout()
        stats_layout.setAlignment(Qt.AlignLeft)
        stats_layout.setSpacing(6)
        # 统计标签样式
        stats_label_style = f"""
            QLabel {{
                background-color: #ffffff;
                border: 1px solid #0cdfe6;
                border-radius: 8px;
                padding: 6px 10px;
                font-size: {self.font_size}pt;
            }}
        """
        # 统计标签对象
        self.label_total = QLabel("📊总数：0")  # 总数标签
        self.label_done = QLabel("✅完成：0")  # 已完成标签
        self.label_undone = QLabel("⏳剩余：0")  # 未完成标签
        self.label_error = QLabel("❗异常：0")  # 异常标签
        # 应用样式
        self.label_total.setStyleSheet(stats_label_style)
        self.label_done.setStyleSheet(stats_label_style)
        self.label_undone.setStyleSheet(stats_label_style)
        self.label_error.setStyleSheet(stats_label_style)
        # 添加到布局
        stats_layout.addWidget(self.label_total)
        stats_layout.addWidget(self.label_done)
        stats_layout.addWidget(self.label_undone)
        stats_layout.addWidget(self.label_error)
        main_layout.addLayout(stats_layout)

        # --- 数据表格 ---
        self.table = QTableWidget()  # 表格控件
        self.table.setAlternatingRowColors(True)  # 开启隔行变色（灰白相间）
        self.table.verticalHeader().setDefaultSectionSize(self.table_row_height)  # 表格行高，根据平台自动调整
        self.table.verticalHeader().setVisible(False)  # 隐藏行号
        # 表格美化样式
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: #ffffff;
                gridline-color: #f0f0f0;
                border: 1px solid #dcdfe6;
                border-radius: 8px;
                alternate-background-color: #f5f7fa;  /* 灰白相间：偶数行浅灰色，奇数行白色 */
            }}
            QHeaderView::section {{
                border: none;
                border-bottom: 1px solid #dcdfe6;
                border-right: 1px solid #f0f0f0;
                background-color: #f5f7fa;
                padding: 10px 8px;
                font-size: {self.font_size}pt;
                color: #606266;
                font-weight: 500;
            }}
            QTableWidget::item {{
                padding: 4px;
                border: none;
                font-size: {self.font_size}pt;
            }}
            QTableWidget::item:selected {{
                background-color: #e5f2ff;
                color: inherit;
            }}
            /* 表格滚动条美化 */
            QScrollBar:vertical {{
                width: 8px;
                background: #f5f7fa;
                border-radius: 4px;
                margin: 2px;
            }}
            QScrollBar::handle:vertical {{
                background: #c0c4cc;
                border-radius: 4px;
                min-height: 20px;
            }}
            QScrollBar:horizontal {{
                height: 8px;
                background: #f5f7fa;
                border-radius: 4px;
                margin: 2px;
            }}
            QScrollBar::handle:horizontal {{
                background: #c0c4cc;
                border-radius: 4px;
                min-width: 20px;
            }}
        """)
        self.table.cellClicked.connect(self.on_cell_clicked)  # 绑定点击事件：点击行弹出异常信息
        main_layout.addWidget(self.table)

        # --- 功能按钮组：筛选、完成、未完成、修改、清空 ---
        btn_layout = QHBoxLayout()
        btn_layout.setAlignment(Qt.AlignLeft)
        btn_layout.setSpacing(10)
        btn_width = self.btn_fixed_width + 5  # 功能按钮稍微大一点

        # 筛选按钮
        self.filter_btn = QPushButton("🔍 筛选")
        self.filter_btn.setFixedWidth(btn_width)
        self.filter_btn.setStyleSheet(f"""
            QPushButton{{
                background-color: #009399;
                color: white;
                border: none;
                border-radius: 6px;
                padding: {self.btn_padding}px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color: #a6a9ad;}}
            QPushButton:pressed{{background-color: #76787c;}}
        """)
        # 完成按钮
        self.btn_done = QPushButton("✅ 完成")
        self.btn_done.setFixedWidth(btn_width)
        self.btn_done.setStyleSheet(f"""
            QPushButton{{
                background-color:#67c23a;
                color:white;
                border:none;
                border-radius:4px;
                padding: {self.btn_padding}px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color:#85ce61;}}
            QPushButton:pressed{{background-color:#529b2e;}}
        """)
        # 未完成按钮
        self.btn_undone = QPushButton("⏳ 未完成")
        self.btn_undone.setFixedWidth(btn_width)
        self.btn_undone.setStyleSheet(f"""
            QPushButton{{
                background-color:#e6a23c;
                color:white;
                border:none;
                border-radius:4px;
                padding: {self.btn_padding}px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color:#ebb563;}}
            QPushButton:pressed{{background-color:#cf8a24;}}
        """)
        # 修改按钮
        self.btn_modify = QPushButton("✏️ 修改")
        self.btn_modify.setFixedWidth(btn_width)
        self.btn_modify.setStyleSheet(f"""
            QPushButton{{
                background-color:#909399;
                color:white;
                border:none;
                border-radius:4px;
                padding: {self.btn_padding}px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color:#a6a9ad;}}
            QPushButton:pressed{{background-color:#76787c;}}
        """)
        # 清空按钮
        self.btn_clear = QPushButton("🗑️ 清空")
        self.btn_clear.setFixedWidth(btn_width)
        self.btn_clear.setStyleSheet(f"""
            QPushButton{{
                background-color:#f56c6c;
                color:white;
                border:none;
                border-radius:4px;
                padding: {self.btn_padding}px;  /* 按钮高度，根据平台自动调整 */
                font-size: {self.font_size}pt;
            }}
            QPushButton:hover{{background-color:#f78989;}}
            QPushButton:pressed{{background-color:#dd5252;}}
        """)

        # 按顺序添加按钮
        btn_layout.addWidget(self.filter_btn)
        btn_layout.addWidget(self.btn_done)
        btn_layout.addWidget(self.btn_undone)
        btn_layout.addWidget(self.btn_modify)
        btn_layout.addWidget(self.btn_clear)
        main_layout.addLayout(btn_layout)

        # ==================== 事件绑定 ====================
        self.filter_btn.clicked.connect(self.start_filter)
        self.import_btn.clicked.connect(self.import_file)
        self.save_btn.clicked.connect(self.save_as_file)
        self.btn_done.clicked.connect(self.set_done)
        self.btn_undone.clicked.connect(self.set_undone)
        self.btn_modify.clicked.connect(self.do_modify)
        self.btn_clear.clicked.connect(self.clear_all)

    # -------------------------------------------------------------------------
    # 安卓返回键处理
    # -------------------------------------------------------------------------
    def keyPressEvent(self, event):
        """
        重写按键事件，处理安卓的返回键
        避免用户误触返回键直接退出程序
        """
        if self.is_android and event.key() == Qt.Key_Back:
            # 安卓返回键，弹出确认退出
            reply = QMessageBox.question(self, "确认退出", "确定要退出上画管理系统吗？",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                event.accept()
                QApplication.quit()
            else:
                event.ignore()
        else:
            super().keyPressEvent(event)

    # -------------------------------------------------------------------------
    # 权限申请回调
    # -------------------------------------------------------------------------
    @Slot(str, bool)
    def on_permission_result(self, permission, granted):
        """
        权限申请的回调函数
        当用户授权/拒绝权限后触发
        """
        if granted and self._pending_import:
            # 权限申请成功，继续执行导入文件
            self._pending_import = False
            self._do_import_file()
        elif not granted:
            # 权限被拒绝，提示用户
            QMessageBox.warning(self, "权限不足", "需要存储权限才能读取Excel文件，请在设置中开启！")

    # -------------------------------------------------------------------------
    # 筛选弹窗相关
    # -------------------------------------------------------------------------
    def pick_item(self, title, items):
        """
        生成选择弹窗（无OK/Cancel，点选即确认）
        :param title: 弹窗标题
        :param items: 选项列表
        :return: 选中的项，取消则返回None
        """
        dlg = QDialog(self)  # 弹窗对象
        dlg.setWindowTitle(title)
        # 弹窗美化样式，根据平台调整
        dlg.setStyleSheet(f"""
            QDialog {{
                background-color: #ffffff;
            }}
            QListWidget {{
                border: 1px solid #dcdfe6;
                border-radius: 6px;
                padding: 4px;
                font-size: {self.font_size}pt;
                outline: none;
            }}
            QListWidget::item {{
                height: {self.popup_item_height}px;  /* 弹窗选项高度，根据平台自动调整 */
                padding: 0px 8px;
                border-radius: 4px;
            }}
            QListWidget::item:selected {{
                background-color: #409eff;
                color: white;
            }}
            QListWidget::item:hover {{
                background-color: #e5f2ff;
            }}
            QScrollBar:vertical {{
                width: 5px;
                background: #f5f7fa;
            }}
            QScrollBar::handle:vertical {{
                background: #c0c4cc;
                border-radius: 2px;
            }}
        """)

        # 列表控件
        list_widget = QListWidget()  # 列表选择控件
        list_widget.addItems(items)

        # 弹窗大小自适应：最多10行，少于10行自动缩小
        item_height = self.popup_item_height
        max_visible = 10  # 最多显示的选项行数
        total_items = len(items)  # 总选项数
        visible = min(total_items, max_visible)  # 实际显示的行数
        win_w = self.popup_width  # 弹窗宽度，根据平台自动调整
        win_h = visible * item_height + 20  # 弹窗高度

        dlg.setFixedSize(win_w, win_h)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.addWidget(list_widget)

        # 点击选项自动关闭
        selected_value = None  # 存储选中的值

        def on_click(item):
            nonlocal selected_value
            selected_value = item.text()
            dlg.close()

        list_widget.itemClicked.connect(on_click)

        # ==================== 弹窗位置 ====================
        if self.is_android:
            # 安卓上弹窗居中，避免被底部导航栏挡住
            dlg.move(self.geometry().center() - dlg.rect().center())
        else:
            # PC上保持原来的底部位置
            main_geo = self.geometry()
            x = main_geo.x() + (main_geo.width() - win_w) / 2
            y = main_geo.y() + main_geo.height() - win_h - 12
            dlg.move(x, y)

        dlg.exec()
        return selected_value

    def start_filter(self):
        """
        开始筛选流程：依次弹窗选择小区→楼栋→单元→电梯→框位
        每选一步，表格实时刷新
        """
        if self.df_original is None:
            QMessageBox.warning(self, "提示", "请先导入文件！")
            return

        # 重置筛选条件
        self.filter_condition = {k: "全部" for k in self.filter_condition}
        self.refresh_all()

        # 1. 选择小区
        area_list = ["全部"] + sorted(self.df_original.iloc[:, 0].dropna().unique().tolist())  # 小区选项列表
        area = self.pick_item("选择小区", area_list)  # 用户选择的小区
        if area is None:
            return
        self.filter_condition["area"] = area
        self.refresh_all()  # 实时刷新表格

        if area != "全部":
            # 2. 选择楼栋
            build_df = self.df_original[self.df_original.iloc[:, 0] == area]  # 筛选后的楼栋数据
            build_list = ["全部"] + sorted(build_df.iloc[:, 1].dropna().unique().tolist())  # 楼栋选项列表
            build = self.pick_item("选择楼栋", build_list)  # 用户选择的楼栋
            if build is None:
                return
            self.filter_condition["build"] = build
            self.refresh_all()  # 实时刷新表格

            if build != "全部":
                # 3. 选择单元
                unit_df = build_df[build_df.iloc[:, 1] == build]  # 筛选后的单元数据
                unit_list = ["全部"] + sorted(unit_df.iloc[:, 2].dropna().unique().tolist())  # 单元选项列表
                unit = self.pick_item("选择单元", unit_list)  # 用户选择的单元
                if unit is None:
                    return
                self.filter_condition["unit"] = unit
                self.refresh_all()  # 实时刷新表格

                if unit != "全部":
                    # 4. 选择电梯
                    elevator_df = unit_df[unit_df.iloc[:, 2] == unit]  # 筛选后的电梯数据
                    elevator_list = ["全部"] + sorted(elevator_df.iloc[:, 3].dropna().unique().tolist())  # 电梯选项列表
                    elevator = self.pick_item("选择电梯", elevator_list)  # 用户选择的电梯
                    if elevator is None:
                        return
                    self.filter_condition["elevator"] = elevator
                    self.refresh_all()  # 实时刷新表格

                    if elevator != "全部":
                        # 5. 选择框位
                        box_df = elevator_df[elevator_df.iloc[:, 3] == elevator]  # 筛选后的框位数据
                        box_list = ["全部"] + sorted(box_df.iloc[:, 4].dropna().unique().tolist())  # 框位选项列表
                        box = self.pick_item("选择框位", box_list)  # 用户选择的框位
                        if box is None:
                            return
                        self.filter_condition["box"] = box
                        self.refresh_all()  # 实时刷新表格

    # -------------------------------------------------------------------------
    # 表格与数据刷新
    # -------------------------------------------------------------------------
    def on_cell_clicked(self, row, col):
        """
        点击表格行，弹出异常信息
        :param row: 点击的行号
        :param col: 点击的列号
        """
        if self.df_display is None:
            return
        try:
            data_row = self.df_display.iloc[row]  # 点击的行数据
            info = str(data_row.iloc[7]).strip()  # 异常信息
            if info and info != "":
                QMessageBox.information(self, "异常信息", info)
        except:
            pass

    def refresh_all(self):
        """
        根据筛选条件刷新表格和统计
        """
        if self.df_original is None:
            return

        # 应用筛选条件
        a = self.filter_condition["area"]  # 筛选条件：小区
        b = self.filter_condition["build"]  # 筛选条件：楼栋
        u = self.filter_condition["unit"]  # 筛选条件：单元
        e = self.filter_condition["elevator"]  # 筛选条件：电梯
        c = self.filter_condition["box"]  # 筛选条件：框位

        df = self.df_original.copy()  # 临时数据副本
        if a != "全部":
            df = df[df.iloc[:, 0] == a]
        if b != "全部":
            df = df[df.iloc[:, 1] == b]
        if u != "全部":
            df = df[df.iloc[:, 2] == u]
        if e != "全部":
            df = df[df.iloc[:, 3] == e]
        if c != "全部":
            df = df[df.iloc[:, 4] == c]

        self.df_display = df  # 保存筛选后的数据
        self.show_table()  # 刷新表格
        self.update_stats()  # 刷新统计

    def show_table(self):
        """
        渲染表格数据，根据行样式设置字体颜色
        """
        if self.df_display is None:
            self.table.setRowCount(0)
            return
        df = self.df_display  # 当前显示的数据
        self.table.setRowCount(len(df))
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(['小区', '楼栋', '单元', '电梯', '框位', '客户', '备注'])
        for i in range(len(df)):
            key = tuple(df.iloc[i, 0:7].astype(str).tolist())  # 行的唯一标识
            color = self.row_styles.get(key, "black")  # 行的颜色样式
            # 颜色映射：绿色=完成，红色=异常，黑色=正常
            qc = QColor(0, 180, 0) if color == "green" else QColor(255, 0, 0) if color == "red" else QColor(0, 0, 0)
            for j in range(7):
                item = QTableWidgetItem(str(df.iloc[i, j]))
                item.setForeground(qc)
                self.table.setItem(i, j, item)
        # 列宽自适应：根据每列内容自动调整宽度，确保所有内容完整显示，不会被截断
        self.table.resizeColumnsToContents()

    def update_stats(self):
        """
        更新统计标签
        """
        if self.df_original is None:
            total = done = error = undone = 0
        else:
            total = len(self.df_original)  # 总任务数
            done = sum(1 for k in self.row_styles if self.row_styles[k] == "green")  # 已完成数
            error = sum(1 for k in self.row_styles if self.row_styles[k] == "red")  # 异常数
            undone = total - done - error  # 未完成数
        self.label_total.setText(f"📊总数：{total}")
        self.label_done.setText(f"✅完成：{done}")
        self.label_undone.setText(f"⏳剩余：{undone}")
        self.label_error.setText(f"❗异常：{error}")

    # -------------------------------------------------------------------------
    # Excel样式读取
    # -------------------------------------------------------------------------
    def read_excel_styles(self):
        """
        读取Excel文件的字体颜色，同步到row_styles
        """
        if not self.file_path:
            return
        wb = openpyxl.load_workbook(self.file_path)  # 打开Excel文件
        ws = wb.active  # 活动工作表
        self.row_styles = {}  # 重置行样式
        for r in range(2, ws.max_row + 1):  # 遍历所有行
            # 生成行的唯一标识
            key = tuple(str(ws.cell(r, c).value or "").strip() for c in range(1, 8))
            color = "black"  # 默认颜色
            cell = ws.cell(r, 1)
            # 读取字体颜色
            if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb'):
                rgb = str(cell.font.color.rgb).upper()
                if "00B400" in rgb:
                    color = "green"
                elif "FF0000" in rgb:
                    color = "red"
            self.row_styles[key] = color
        wb.close()

    # -------------------------------------------------------------------------
    # 导入导出
    # -------------------------------------------------------------------------
    def _do_import_file(self):
        """
        实际执行导入文件的逻辑，权限申请成功后调用
        """
        # 安卓上默认打开下载目录
        path, _ = QFileDialog.getOpenFileName(
            filter="Excel Files (*.xlsx)",
            dir=self.default_file_dir
        )  # 选择文件
        if not path:
            return
        self.file_path = path  # 保存文件路径
        # 读取数据
        df = pd.read_excel(path, nrows=1000)  # 读取Excel数据
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())  # 数据清洗：空值转空字符串，去空格
        self.df_original = df  # 保存原始数据
        self.df_display = df.copy()  # 显示数据初始化为全部
        # 读取样式
        self.read_excel_styles()
        self.refresh_all()  # 刷新界面

    def import_file(self):
        """
        导入Excel文件，安卓上先申请权限
        """
        if self.is_android:
            # 安卓平台：先申请存储权限
            permission = "android.permission.READ_EXTERNAL_STORAGE"
            if not QtAndroid.checkPermission(permission):
                # 没有权限，申请权限
                self._pending_import = True
                QtAndroid.requestPermission(permission, self.on_permission_result)
                return
            # 有权限，直接执行
        # PC平台直接执行
        self._do_import_file()

    def save_as_file(self):
        """
        另存为Excel文件
        """
        if self.df_original is None:
            QMessageBox.warning(self, "提示", "请先导入文件！")
            return
        # 安卓上默认保存到下载目录
        save_path, _ = QFileDialog.getSaveFileName(
            filter="Excel Files (*.xlsx)",
            dir=self.default_file_dir
        )  # 保存路径
        if not save_path:
            return
        if not save_path.endswith(".xlsx"):
            save_path += ".xlsx"
        self.df_original.to_excel(save_path, index=False)
        QMessageBox.information(self, "成功", "文件已保存到下载目录！")

    # -------------------------------------------------------------------------
    # 状态修改：完成/未完成
    # -------------------------------------------------------------------------
    def set_done(self):
        """
        将筛选结果标记为完成（绿色字体）
        """
        if self.df_display is None:
            return
        has_all = any(v == "全部" for v in self.filter_condition.values())  # 是否有筛选条件为全部
        if has_all:
            if QMessageBox.question(self, "确认", "确定要修改全部？") != QMessageBox.Yes:
                return
        # 修改Excel字体颜色
        wb = openpyxl.load_workbook(self.file_path)  # 打开Excel
        ws = wb.active  # 工作表
        ft = openpyxl.styles.Font(color="00B400")  # 绿色字体
        for i in range(len(self.df_display)):
            key = tuple(self.df_display.iloc[i, 0:7].astype(str).tolist())  # 行的唯一标识
            self.row_styles[key] = "green"
            for r in range(2, ws.max_row + 1):
                if all(str(ws.cell(r, x + 1).value or "").strip() == str(self.df_display.iloc[i, x]) for x in range(7)):
                    for cc in range(1, 8):
                        ws.cell(r, cc).font = ft
                    break
        wb.save(self.file_path)
        wb.close()

        # 修复BUG：重新加载数据，同步Excel修改，避免旧异常信息残留
        df = pd.read_excel(self.file_path, nrows=1000)
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        self.df_original = df
        self.df_display = df.copy()
        self.read_excel_styles()
        self.refresh_all()

    def set_undone(self):
        """
        将筛选结果标记为未完成（黑色字体，清空异常信息）
        """
        if self.df_display is None:
            return
        has_all = any(v == "全部" for v in self.filter_condition.values())  # 是否有筛选条件为全部
        if has_all:
            if QMessageBox.question(self, "确认", "确定要修改全部？") != QMessageBox.Yes:
                return

        # 修改Excel：清空异常列，字体改黑色
        wb = openpyxl.load_workbook(self.file_path)  # 打开Excel
        ws = wb.active  # 工作表
        ft = openpyxl.styles.Font(color="000000")  # 黑色字体

        for i in range(len(self.df_display)):
            row_data = self.df_display.iloc[i]
            for r in range(2, ws.max_row + 1):
                if all(str(ws.cell(r, c + 1).value or "").strip() == str(row_data.iloc[c]).strip() for c in range(7)):
                    ws.cell(r, 8, "")  # 清空异常信息
                    for cc in range(1, 8):
                        ws.cell(r, cc).font = ft
                    break

        wb.save(self.file_path)
        wb.close()

        # 修复BUG：重新加载数据，同步Excel修改，避免旧异常信息残留
        df = pd.read_excel(self.file_path, nrows=1000)
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        self.df_original = df
        self.df_display = df.copy()
        self.read_excel_styles()
        self.refresh_all()

    # -------------------------------------------------------------------------
    # 数据修改
    # -------------------------------------------------------------------------
    def do_modify(self):
        """
        修改框位信息：只有筛选结果为1行时才能修改
        """
        # 判断：筛选结果必须为1行
        if self.df_display is None or len(self.df_display) != 1:
            QMessageBox.warning(self, "提示", "请选择具体框位（筛选结果必须为1行）！")
            return

        # 获取当前行数据
        row = self.df_display.iloc[0]
        area = str(row.iloc[0]).strip()  # 原小区
        build = str(row.iloc[1]).strip()  # 原楼栋
        unit = str(row.iloc[2]).strip()  # 原单元
        elevator = str(row.iloc[3]).strip()  # 原电梯
        box = str(row.iloc[4]).strip()  # 原框位

        # 输入修改原因
        reason, ok = QInputDialog.getText(self, "修改", "请输入修改原因：")
        if not ok or not reason:
            return

        # 输入新的框位信息
        while True:
            new_area, ok_a = QInputDialog.getText(self, "修改", "新小区：", text=area)
            new_build, ok_b = QInputDialog.getText(self, "修改", "新楼栋：", text=build)
            new_unit, ok_u = QInputDialog.getText(self, "修改", "新单元：", text=unit)
            new_elev, ok_e = QInputDialog.getText(self, "修改", "新电梯：", text=elevator)
            new_box, ok_bx = QInputDialog.getText(self, "修改", "新框位：", text=box)
            if not all([ok_a, ok_b, ok_u, ok_e, ok_bx]):
                return

            # 检查新框位是否已存在
            exist = self.df_original[
                (self.df_original.iloc[:, 0] == new_area) &
                (self.df_original.iloc[:, 1] == new_build) &
                (self.df_original.iloc[:, 2] == new_unit) &
                (self.df_original.iloc[:, 3] == new_elev) &
                (self.df_original.iloc[:, 4] == new_box)
                ]
            if len(exist) > 0:
                QMessageBox.warning(self, "冲突", "框位已存在，请重新输入！")
                continue
            break

        # 生成异常信息
        origin_str = f"{area}:{build}:{unit}:{elevator}:{box}"
        exception_str = f"原位置[{origin_str}] → 修改原因：{reason}"

        # 修改Excel
        wb = openpyxl.load_workbook(self.file_path)  # 打开Excel
        ws = wb.active  # 工作表
        font_red = openpyxl.styles.Font(color="FF0000")  # 红色字体

        for r in range(2, ws.max_row + 1):
            a = str(ws.cell(r, 1).value or "").strip()
            b = str(ws.cell(r, 2).value or "").strip()
            u = str(ws.cell(r, 3).value or "").strip()
            e = str(ws.cell(r, 4).value or "").strip()
            bx = str(ws.cell(r, 5).value or "").strip()

            if a == area and b == build and u == unit and e == elevator and bx == box:
                ws.cell(r, 1, new_area)
                ws.cell(r, 2, new_build)
                ws.cell(r, 3, new_unit)
                ws.cell(r, 4, new_elev)
                ws.cell(r, 5, new_box)
                ws.cell(r, 8, exception_str)
                for col in range(1, 8):
                    ws.cell(r, col).font = font_red
                break

        wb.save(self.file_path)
        wb.close()

        # 重新加载数据
        df = pd.read_excel(self.file_path, nrows=1000)
        df = df.fillna("").astype(str)
        self.df_original = df
        self.df_display = df.copy()
        self.read_excel_styles()

        # 自动筛选到新位置
        self.filter_condition = {
            "area": new_area,
            "build": new_build,
            "unit": new_unit,
            "elevator": new_elev,
            "box": new_box
        }
        self.refresh_all()
        QMessageBox.information(self, "成功", "修改完成！")

    # -------------------------------------------------------------------------
    # 清空筛选
    # -------------------------------------------------------------------------
    def clear_all(self):
        """
        清空筛选条件，显示全部数据
        """
        if not self.file_path:
            return
        # 重新加载数据
        df = pd.read_excel(self.file_path, nrows=1000)
        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        self.df_original = df
        self.df_display = df.copy()
        self.read_excel_styles()
        # 重置筛选条件
        self.filter_condition = {k: "全部" for k in self.filter_condition}
        self.refresh_all()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = InfoManagerWindow()
    win.show()
    sys.exit(app.exec())
